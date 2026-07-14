import json
import os
import sys
import threading
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Literal

import openpyxl
from pydantic import BaseModel, field_validator

from ask import resolve_cf
from auth import authorize, log_audit
from cli_session import read_session
from dental_notes_schema import CF_PATTERN, DentalNote
from extract_note import OllamaUnreachable, extract_json
from storage import get_collection, init_db, lookup_patient, upsert_note_chroma, upsert_note_sql

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL_ID = "llama3.2:3b"  # general model prompted to emit tool JSON, not dental-notes

DB_PATH = "db/clinic.sqlite"
UNDO_LOG = "db/undo_log.jsonl"

# serialize every read-modify-write of the undo log - the web app serves
# threaded, so an append landing between undo_last's read and its rewrite
# would otherwise be silently dropped
_undo_lock = threading.Lock()

# editable field -> sqlite column; the column is always taken from here,
# never from model output
EDITABLE_FIELDS = {"phone": "phone", "patient_name": "patient_name"}
INVOICE_HEADER = ["date", "amount", "description"]

INTERPRETER_PROMPT = (
    "You are a clinic assistant that turns a command into ONE tool call.\n"
    "Reply with exactly one JSON object and nothing else: "
    '{"tool": "<tool name>", "args": {...}}\n'
    "Tools:\n"
    '- update_field: args {"patient": str, "field": str, "value": str}\n'
    '- append_note: args {"patient": str, "text": str}\n'
    '- add_invoice: args {"patient": str, "amount": float, "description": str}\n'
    "Rules:\n"
    "- use add_invoice whenever the command mentions an invoice or an amount of money\n"
    "- use append_note whenever the command says to append or add to a note\n"
    "- use update_field only when the command sets a field like phone to a new value\n"
    "- patient is always the person's name, never a treatment\n"
    "- text is the exact words after the colon, copied verbatim\n"
    "- description is the treatment or service, never a name\n"
    "- take every value from the command itself, never invent one\n"
    "Command: "
)


class UpdateFieldArgs(BaseModel):
    patient: str
    field: str
    value: str

    @field_validator("field")
    @classmethod
    def validate_field(cls, v):
        if v not in EDITABLE_FIELDS:
            raise ValueError(f"field must be one of {sorted(EDITABLE_FIELDS)}, got {v!r}")
        return v


class AppendNoteArgs(BaseModel):
    patient: str
    text: str


class AddInvoiceArgs(BaseModel):
    patient: str
    amount: float
    description: str


TOOL_ARGS = {
    "update_field": UpdateFieldArgs,
    "append_note": AppendNoteArgs,
    "add_invoice": AddInvoiceArgs,
}


class ToolCall(BaseModel):
    tool: Literal["update_field", "append_note", "add_invoice"]
    args: dict

    def parsed_args(self):
        return TOOL_ARGS[self.tool](**self.args)


def parse_tool_call(reply):
    # reply -> validated ToolCall with typed args. Raises ValueError so a
    # malformed, unknown, or non-whitelisted tool call never gets executed.
    obj = extract_json(reply)
    if obj is None:
        raise ValueError("model did not return valid JSON")
    try:
        call = ToolCall(**obj)
        call.parsed_args()
    except Exception as e:
        raise ValueError("model output failed schema validation: " + str(e))
    return call


def call_model(command, urlopen=urllib.request.urlopen):
    payload = {
        "model": MODEL_ID,
        "prompt": INTERPRETER_PROMPT + command,
        "stream": False,
        "format": "json",
        "options": {"temperature": 0},
    }
    data = json.dumps(payload).encode()
    req = urllib.request.Request(OLLAMA_URL, data=data, headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=120) as resp:
            body = json.load(resp)
    except (urllib.error.URLError, json.JSONDecodeError, TimeoutError):
        # a refused connection, a 200 with a non-JSON body (mid-restart, proxy
        # page), or a read timeout all mean "Ollama isn't usable right now"
        raise OllamaUnreachable("Ollama not reachable - run: ollama run llama3.2:3b")
    return body.get("response", "")


def write_undo_entry(entry, log_path=UNDO_LOG):
    Path(log_path).parent.mkdir(parents=True, exist_ok=True)
    with _undo_lock:
        with open(log_path, "a") as f:
            f.write(json.dumps(entry) + "\n")


def _rewrite_log(log_file, lines):
    # crash-safe replace: write the surviving lines to a sibling temp file and
    # os.replace it over the log, so a crash mid-write can't truncate history
    text = "\n".join(lines) + ("\n" if lines else "")
    tmp = log_file.with_suffix(log_file.suffix + ".tmp")
    tmp.write_text(text)
    os.replace(tmp, log_file)


def ask_cf(candidates):
    # cli-only chooser - never reachable from the web routes, which pass
    # their own choose_cf into build_pending_action
    print("candidates: " + ", ".join(candidates))
    return input("type the codice fiscale to use: ").strip().upper()


def resolve_patient(name, conn, choose_cf=None):
    # returns (cf, reason): reason is None on success, else a specific message
    # the web can show instead of a generic "check the name / permissions"
    cf = resolve_cf(name, conn)
    if cf is None:
        return None, f"no patient named {name} on record"
    if isinstance(cf, list):
        if choose_cf is None:
            # non-interactive caller (web) - refuse rather than prompt
            return None, f"multiple patients named {name} found - be more specific"
        typed = choose_cf(cf)
        # candidates are already CF_PATTERN-filtered in resolve_cf, so
        # membership also guarantees a pattern-valid cf
        if typed not in cf:
            return None, "that codice fiscale is not one of the candidates"
        cf = typed
    return cf, None


def build_diff_line(field, current, new, name, cf):
    return f"{field}: {current} -> {new} ({name}, {cf})"


def confirm(input_fn=input):
    answer = input_fn("proceed? [y/N]: ").strip().lower()
    return answer in ("y", "yes")


def update_field(cf, field, value, conn):
    column = EDITABLE_FIELDS[field]
    conn.execute(f"UPDATE patients SET {column} = ? WHERE codice_fiscale = ?", (value, cf))
    conn.commit()


def add_invoice(cf, amount, description, visit_date, sorted_root=Path("sorted")):
    # cf must be validated before any path is built - a model-supplied value
    # containing "../" must never reach the filesystem (T-06-02).
    if not CF_PATTERN.match(cf):
        raise ValueError(f"codice_fiscale must match ^[A-Z]{{4}}[0-9]{{12}}$, got {cf!r}")

    xlsx_dir = sorted_root / cf / "records"
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = xlsx_dir / "invoices.xlsx"

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        row_count = ws.max_row - 1  # rows before this append, excluding header
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(INVOICE_HEADER)
        row_count = 0

    ws.append([visit_date, amount, description])
    wb.save(xlsx_path)
    return row_count


def pick_target_visit(cf, conn):
    # most recent visit for the CF (D-11 discretion); also returns the total
    # visit count so the confirm diff can say "most recent of N"
    # newest by visit_date, not by insert order - a bulk load_from_sorted
    # re-import assigns ids by filename, so the highest id is not the latest
    # visit; nulls sort last so a dated visit always wins over an undated one
    row = conn.execute(
        "SELECT source_path, clinical_notes, visit_date FROM visits"
        " WHERE codice_fiscale = ? ORDER BY visit_date IS NULL, visit_date DESC, id DESC LIMIT 1",
        (cf,),
    ).fetchone()
    if row is None:
        return None
    count = conn.execute(
        "SELECT COUNT(*) c FROM visits WHERE codice_fiscale = ?", (cf,)
    ).fetchone()["c"]
    return row["source_path"], row["clinical_notes"], row["visit_date"], count


def append_note(cf, text, source_path, conn, collection, sorted_root=Path("sorted")):
    # cf must be validated before any path is built (T-06-02)
    if not CF_PATTERN.match(cf):
        raise ValueError(f"codice_fiscale must match ^[A-Z]{{4}}[0-9]{{12}}$, got {cf!r}")

    json_path = sorted_root / cf / "notes" / (Path(source_path).stem + ".json")
    note = DentalNote.model_validate_json(json_path.read_text())
    note.clinical_notes = (note.clinical_notes + "\n" + text) if note.clinical_notes else text
    json_path.write_text(note.model_dump_json())

    upsert_note_sql(note, source_path, conn)
    upsert_note_chroma(note, source_path, collection)


def build_pending_action(call, conn, role, username, sorted_root=Path("sorted"), choose_cf=None):
    # returns (pending, reason): pending is the frozen-payload dict on success,
    # else (None, reason) with a specific message. authorize check #1 here is
    # an early UX denial - apply_pending_action re-checks it independently with
    # the live role at apply time (SC4).
    args = call.parsed_args()

    if not authorize(role, call.tool):
        log_audit(conn, username, role, call.tool, target=args.patient, allowed=0)
        return None, f"not permitted: {role} may not {call.tool}"

    cf, reason = resolve_patient(args.patient, conn, choose_cf)
    if cf is None:
        return None, reason

    if call.tool == "update_field":
        data = lookup_patient(cf, conn)
        current = data[EDITABLE_FIELDS[args.field]]
        name = data["patient_name"]
        diff_line = build_diff_line(args.field, current, args.value, name, cf)
        return {
            "tool": call.tool,
            "args": {"field": args.field, "value": args.value},
            "cf": cf,
            "diff_line": diff_line,
            "before": current,
            "target": f"sqlite:patients.{args.field}",
        }, None

    elif call.tool == "append_note":
        target = pick_target_visit(cf, conn)
        if target is None:
            return None, f"no visit on record for {cf}"
        source_path, current_notes, visit_date, count = target
        # the sqlite row and the json sibling are separately mutable - bail
        # before the pending action is built for an edit that can't happen
        json_path = sorted_root / cf / "notes" / (Path(source_path).stem + ".json")
        if not json_path.exists():
            return None, f"note file missing for this visit ({json_path}) - fix the sorted tree first"
        diff_line = f"appending to visit from {visit_date} (most recent of {count})"
        return {
            "tool": call.tool,
            "args": {"text": args.text},
            "cf": cf,
            "diff_line": diff_line,
            "before": current_notes,
            "target": f"visit:{source_path}",
        }, None

    elif call.tool == "add_invoice":
        data = lookup_patient(cf, conn)
        visit_date = datetime.now().date().isoformat()
        diff_line = (f"add invoice row for {data['patient_name']} ({cf}): "
                     f"{visit_date} | {args.amount} | {args.description}")
        xlsx_path = sorted_root / cf / "records" / "invoices.xlsx"
        before = 0
        if xlsx_path.exists():
            before = openpyxl.load_workbook(xlsx_path).active.max_row - 1
        return {
            "tool": call.tool,
            "args": {"amount": args.amount, "description": args.description, "visit_date": visit_date},
            "cf": cf,
            "diff_line": diff_line,
            "before": before,
            "target": f"xlsx:{xlsx_path}",
        }, None


def apply_pending_action(pending, conn, role, username, log_path=UNDO_LOG,
                          collection=None, sorted_root=Path("sorted")):
    # authorize check #2 - SC4's actual security boundary. Uses the live role
    # passed in from the current request, never pending['role'] (there is no
    # such key) - a forged/replayed apply call must still be denied here,
    # independent of whatever build_pending_action already checked.
    if not authorize(role, pending["tool"]):
        log_audit(conn, username, role, pending["tool"], target=pending["cf"], allowed=0)
        print(f"not permitted: {role} may not {pending['tool']}")
        return

    # mutate first - if the apply blows up (e.g. the note file vanished
    # during the confirm window) there must be no phantom undo entry or
    # audit row claiming a change that never happened
    args = pending["args"]
    if pending["tool"] == "update_field":
        update_field(pending["cf"], args["field"], args["value"], conn)
    elif pending["tool"] == "append_note":
        source_path = pending["target"][len("visit:"):]
        append_note(pending["cf"], args["text"], source_path, conn, collection, sorted_root)
    elif pending["tool"] == "add_invoice":
        add_invoice(pending["cf"], args["amount"], args["description"], args["visit_date"], sorted_root)

    write_undo_entry({
        "ts": datetime.now().isoformat(),
        "tool": pending["tool"],
        "codice_fiscale": pending["cf"],
        "target": pending["target"],
        "before": pending["before"],
        "username": username,
    }, log_path)
    log_audit(conn, username, role, pending["tool"], target=pending["cf"], allowed=1)


def run_command(command, conn, dry_run, urlopen, role, username, input_fn=input, log_path=UNDO_LOG,
                 collection=None, sorted_root=Path("sorted"), choose_cf=ask_cf):
    reply = call_model(command, urlopen)
    call = parse_tool_call(reply)

    pending, reason = build_pending_action(call, conn, role, username, sorted_root=sorted_root,
                                            choose_cf=choose_cf)
    if pending is None:
        print(reason)  # denied or unresolvable patient - the denial is already audited
        return

    print(pending["diff_line"])
    if dry_run:
        return

    if not confirm(input_fn):
        print("no changes made")
        return

    apply_pending_action(pending, conn, role, username, log_path=log_path,
                          collection=collection, sorted_root=sorted_root)


def undo_last(conn, role, username, log_path=UNDO_LOG, collection=None, sorted_root=Path("sorted")):
    # returns (status, message): status is one of "empty", "error", "denied",
    # "restored", "manual" so callers can flash/print without re-deciding -
    # "manual" in particular means the entry was consumed but nothing was
    # actually reverted, so the web must not report it as a success
    log_file = Path(log_path)
    with _undo_lock:
        if not log_file.exists():
            return "empty", "nothing to undo"
        lines = log_file.read_text().strip().splitlines()
        if not lines:
            return "empty", "nothing to undo"

        # scan backward for the most recent entry this user made - entries
        # written before the username migration have no "username" key and are
        # skipped, never attributed to whoever happens to call undo next (D-06);
        # a corrupt/truncated line (crash mid-append) is skipped, not fatal
        match_index = None
        for i in range(len(lines) - 1, -1, -1):
            try:
                candidate = json.loads(lines[i])
            except json.JSONDecodeError:
                continue
            if candidate.get("username") == username:
                match_index = i
                entry = candidate
                break
        if match_index is None:
            return "empty", "nothing to undo"

        target = entry["target"]
        cf = entry["codice_fiscale"]

        # figure out which action this restore needs before touching anything -
        # an unknown or invalid target is a graceful no-op, never authorized
        if target.startswith("sqlite:patients."):
            field = target[len("sqlite:patients."):]
            if field not in EDITABLE_FIELDS:
                return "error", f"don't know how to undo target {target!r}"
            action = "update_field"
        elif target.startswith("visit:"):
            # undoing an append rewrites existing note content - that's an edit,
            # not a fresh append, so an assistant (append-only, D-04) can't do it
            action = "edit_note"
        elif target.startswith("xlsx:"):
            action = "add_invoice"
        else:
            return "error", f"don't know how to undo target {target!r}"

        if not authorize(role, action):
            log_audit(conn, username, role, action, target=cf, allowed=0)
            return "denied", f"not permitted: {role} may not undo {action}"
        log_audit(conn, username, role, action, target=cf, allowed=1)

        if action == "update_field":
            update_field(cf, field, entry["before"], conn)
            status, message = "restored", f"restored {field} to {entry['before']} for {cf}"
        elif action == "edit_note":
            source_path = target[len("visit:"):]
            json_path = sorted_root / cf / "notes" / (Path(source_path).stem + ".json")
            note = DentalNote.model_validate_json(json_path.read_text())
            note.clinical_notes = entry["before"]
            json_path.write_text(note.model_dump_json())
            upsert_note_sql(note, source_path, conn)
            upsert_note_chroma(note, source_path, collection)
            status, message = "restored", f"restored clinical note text for {cf}"
        elif action == "add_invoice":
            # D-09: invoice rows are restored by hand - drop the entry so the
            # edits beneath it stay reachable, but report "manual" so no caller
            # claims the row was reverted
            xlsx_path = target[len("xlsx:"):]
            status = "manual"
            message = f"cannot auto-undo an invoice row append at {xlsx_path} - restore manually"

        # remove only this entry, keep every other line (including other users'
        # more-recent entries) in place and in order
        rest = lines[:match_index] + lines[match_index + 1:]
        _rewrite_log(log_file, rest)

    return status, message


def selftest():
    import tempfile
    from datetime import date

    with tempfile.TemporaryDirectory() as tmp:
        db_path = str(Path(tmp) / "clinic.sqlite")
        log_path = str(Path(tmp) / "undo_log.jsonl")

        conn = init_db(db_path)
        cf = "RSSM800010150100"
        conn.execute(
            "INSERT INTO patients (codice_fiscale, patient_name, phone) VALUES (?, ?, ?)",
            (cf, "mario rossi", "333 9999999"),
        )
        conn.commit()

        def fake_urlopen(req, timeout=120):
            class FakeResponse:
                def __enter__(self):
                    return self

                def __exit__(self, *exc_info):
                    return False

                def read(self):
                    tool_call = {
                        "tool": "update_field",
                        "args": {"patient": "rossi", "field": "phone", "value": "333-1234"},
                    }
                    return json.dumps({"response": json.dumps(tool_call)}).encode()

            return FakeResponse()

        # a. --dry-run leaves the phone unchanged and writes no log line
        run_command("update rossi's phone to 333-1234", conn, True, fake_urlopen,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333 9999999", "dry-run must not write"
        assert not Path(log_path).exists(), "dry-run must not touch the undo log"

        # b. a declined confirm writes nothing
        run_command("update rossi's phone to 333-1234", conn, False, fake_urlopen,
                     role="dentist", username="test-dentist",
                     input_fn=lambda p: "n", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333 9999999", "declined confirm must not write"
        assert not Path(log_path).exists(), "declined confirm must not write the undo log"

        # c. a confirmed run writes the new phone and exactly one undo-log line
        run_command("update rossi's phone to 333-1234", conn, False, fake_urlopen,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333-1234", "confirmed run must write the new value"
        lines = Path(log_path).read_text().strip().splitlines()
        assert len(lines) == 1, f"expected 1 undo-log line, got {len(lines)}"

        allowed_rows = conn.execute(
            "SELECT * FROM audit_log WHERE action = 'update_field' AND allowed = 1"
        ).fetchall()
        assert len(allowed_rows) == 1, \
            f"confirmed dentist write should log exactly 1 allowed row, got {len(allowed_rows)}"
        assert allowed_rows[0]["username"] == "test-dentist", "allowed row should carry the acting username"
        assert allowed_rows[0]["target"] == cf, "allowed row target should be the cf"

        # c2. an unauthorized role is denied: phone stays unchanged, no new undo
        # line, and exactly one denied audit_log row is recorded (RBAC-05, AUDIT-01)
        lines_before_denial = len(Path(log_path).read_text().strip().splitlines())
        run_command("update rossi's phone to 333-1234", conn, False, fake_urlopen,
                    role="assistant", username="test-assistant",
                    input_fn=lambda p: "y", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333-1234", "denied role must not change the phone"
        assert len(Path(log_path).read_text().strip().splitlines()) == lines_before_denial, \
            "denied role must not add an undo-log line"
        denied_rows = conn.execute(
            "SELECT * FROM audit_log WHERE action = 'update_field' AND allowed = 0"
        ).fetchall()
        assert len(denied_rows) == 1, f"denied update_field should log exactly 1 row, got {len(denied_rows)}"

        assert conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE allowed = 1"
        ).fetchone()["c"] >= 1, "expected at least one allowed=1 audit row"
        assert conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE allowed = 0"
        ).fetchone()["c"] >= 1, "expected at least one allowed=0 audit row"

        # d. undo restores the before-image
        undo_last(conn, role="dentist", username="test-dentist", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333 9999999", "undo must restore the original phone"

        # d2. the undo command is itself a gated write path: an unauthorized
        # role is denied (no mutation, entry not consumed) and an authorized
        # role restores the value and consumes the entry (T-07-14). Same
        # acting user throughout - this checks the authorize gate, not the
        # username-scoped lookup (that's covered separately below).
        write_undo_entry({
            "ts": datetime.now().isoformat(),
            "tool": "update_field",
            "codice_fiscale": cf,
            "target": "sqlite:patients.phone",
            "before": "555-0000",
            "username": "test-dentist",
        }, log_path)

        denied_before = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'update_field' AND allowed = 0"
        ).fetchone()["c"]
        undo_last(conn, role="admin", username="test-dentist", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "333 9999999", "denied undo must not restore the phone"
        assert len(Path(log_path).read_text().strip().splitlines()) == 1, \
            "denied undo must leave the undo-log entry intact"
        denied_after = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'update_field' AND allowed = 0"
        ).fetchone()["c"]
        assert denied_after == denied_before + 1, "denied undo should add exactly one denied audit row"

        allowed_before = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'update_field' AND allowed = 1"
        ).fetchone()["c"]
        undo_last(conn, role="dentist", username="test-dentist", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "555-0000", "allowed undo must restore the before-value"
        assert Path(log_path).read_text().strip() == "", "allowed undo must consume the undo-log entry"
        last_allowed = conn.execute(
            "SELECT * FROM audit_log WHERE action = 'update_field' AND allowed = 1 ORDER BY id DESC LIMIT 1"
        ).fetchone()
        allowed_after = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'update_field' AND allowed = 1"
        ).fetchone()["c"]
        assert allowed_after == allowed_before + 1, "allowed undo should add exactly one allowed audit row"
        assert last_allowed["username"] == "test-dentist", "allowed undo row should carry the acting username"
        assert last_allowed["target"] == cf, "allowed undo row target should be the cf"

        # e. invalid tool JSON and unreachable Ollama are rejected cleanly
        try:
            parse_tool_call("not json")
            raise AssertionError("non-JSON reply should have been rejected")
        except ValueError:
            pass

        def boom(*a, **k):
            raise urllib.error.URLError("connection refused")

        try:
            call_model("anything", urlopen=boom)
            raise AssertionError("unreachable Ollama should raise OllamaUnreachable")
        except OllamaUnreachable:
            pass

        # --- append_note and add_invoice fixtures: a visit with its json sibling
        # and chroma chunk already present, same as a real sort_files/storage load ---
        sorted_root = Path(tmp) / "sorted"
        collection = get_collection(str(Path(tmp) / "chroma"))
        note = DentalNote(
            patient_name="mario rossi",
            codice_fiscale=cf,
            phone="333 9999999",
            visit_date=date(2026, 6, 1),
            procedures=["cleaning"],
            clinical_notes="initial note",
        )
        source_path = f"{cf}/notes/n1.txt"
        notes_dir = sorted_root / cf / "notes"
        notes_dir.mkdir(parents=True)
        (notes_dir / "n1.json").write_text(note.model_dump_json())
        upsert_note_sql(note, source_path, conn)
        upsert_note_chroma(note, source_path, collection)

        def fake_urlopen_append(req, timeout=120):
            class FakeResponse:
                def __enter__(self):
                    return self

                def __exit__(self, *exc_info):
                    return False

                def read(self):
                    tool_call = {
                        "tool": "append_note",
                        "args": {"patient": "rossi", "text": "follow-up done"},
                    }
                    return json.dumps({"response": json.dumps(tool_call)}).encode()

            return FakeResponse()

        # f. append_note dry-run leaves sqlite/json/chroma unchanged and adds no undo line
        lines_before = len(Path(log_path).read_text().strip().splitlines())
        run_command("add a note that follow-up done for rossi", conn, True, fake_urlopen_append,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path,
                    collection=collection, sorted_root=sorted_root)
        row = conn.execute(
            "SELECT clinical_notes FROM visits WHERE source_path = ?", (source_path,)
        ).fetchone()
        assert row["clinical_notes"] == "initial note", "append dry-run must not write sqlite"
        jf = DentalNote.model_validate_json((notes_dir / "n1.json").read_text())
        assert jf.clinical_notes == "initial note", "append dry-run must not write json"
        assert len(Path(log_path).read_text().strip().splitlines()) == lines_before, \
            "append dry-run must not add an undo line"

        # g. a confirmed append updates sqlite + json + chroma and adds one undo line
        run_command("add a note that follow-up done for rossi", conn, False, fake_urlopen_append,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path,
                    collection=collection, sorted_root=sorted_root)
        row = conn.execute(
            "SELECT clinical_notes FROM visits WHERE source_path = ?", (source_path,)
        ).fetchone()
        assert "follow-up done" in row["clinical_notes"], "confirmed append must update sqlite"
        jf = DentalNote.model_validate_json((notes_dir / "n1.json").read_text())
        assert "follow-up done" in jf.clinical_notes, "confirmed append must update json"
        assert collection.count() == 1, "append must not add a second chroma chunk"
        assert len(Path(log_path).read_text().strip().splitlines()) == lines_before + 1, \
            "confirmed append must add exactly one undo line"

        # h. undo restores the original clinical note text through the same sync
        undo_last(conn, role="dentist", username="test-dentist", log_path=log_path,
                  collection=collection, sorted_root=sorted_root)
        row = conn.execute(
            "SELECT clinical_notes FROM visits WHERE source_path = ?", (source_path,)
        ).fetchone()
        assert row["clinical_notes"] == "initial note", "undo must restore the sqlite note text"
        jf = DentalNote.model_validate_json((notes_dir / "n1.json").read_text())
        assert jf.clinical_notes == "initial note", "undo must restore the json note text"

        # h2. a visit undo is gated on edit_note, not append_note - an
        # assistant (append-only, D-04) is denied; a dentist is allowed.
        # Same acting user throughout - this checks the authorize gate, not
        # the username-scoped lookup (that's covered separately below).
        write_undo_entry({
            "ts": datetime.now().isoformat(),
            "tool": "append_note",
            "codice_fiscale": cf,
            "target": f"visit:{source_path}",
            "before": "placeholder before-text",
            "username": "test-dentist",
        }, log_path)

        undo_last(conn, role="assistant", username="test-dentist", log_path=log_path,
                  collection=collection, sorted_root=sorted_root)
        row = conn.execute(
            "SELECT clinical_notes FROM visits WHERE source_path = ?", (source_path,)
        ).fetchone()
        assert row["clinical_notes"] == "initial note", "assistant undo must not touch the note text"
        assert len(Path(log_path).read_text().strip().splitlines()) == 1, \
            "denied visit undo must leave the undo-log entry intact"
        assert conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'edit_note' AND allowed = 0"
        ).fetchone()["c"] == 1, "denied visit undo should log exactly 1 denied edit_note row"

        allowed_edit_before = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'edit_note' AND allowed = 1"
        ).fetchone()["c"]
        undo_last(conn, role="dentist", username="test-dentist", log_path=log_path,
                  collection=collection, sorted_root=sorted_root)
        row = conn.execute(
            "SELECT clinical_notes FROM visits WHERE source_path = ?", (source_path,)
        ).fetchone()
        assert row["clinical_notes"] == "placeholder before-text", "dentist undo must restore the note text"
        assert Path(log_path).read_text().strip() == "", "allowed visit undo must consume the undo-log entry"
        allowed_edit_after = conn.execute(
            "SELECT COUNT(*) c FROM audit_log WHERE action = 'edit_note' AND allowed = 1"
        ).fetchone()["c"]
        assert allowed_edit_after == allowed_edit_before + 1, \
            "allowed visit undo should add exactly one allowed edit_note row"

        def fake_urlopen_invoice(req, timeout=120):
            class FakeResponse:
                def __enter__(self):
                    return self

                def __exit__(self, *exc_info):
                    return False

                def read(self):
                    tool_call = {
                        "tool": "add_invoice",
                        "args": {"patient": "rossi", "amount": 80.0, "description": "cleaning"},
                    }
                    return json.dumps({"response": json.dumps(tool_call)}).encode()

            return FakeResponse()

        xlsx_path = sorted_root / cf / "records" / "invoices.xlsx"

        # i. add_invoice dry-run creates no xlsx file
        run_command("add an invoice for rossi for cleaning 80", conn, True, fake_urlopen_invoice,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path,
                    collection=collection, sorted_root=sorted_root)
        assert not xlsx_path.exists(), "add_invoice dry-run must not create the xlsx file"

        # j. a confirmed add_invoice creates invoices.xlsx with a header row and one data row
        lines_before = len(Path(log_path).read_text().strip().splitlines())
        run_command("add an invoice for rossi for cleaning 80", conn, False, fake_urlopen_invoice,
                    role="dentist", username="test-dentist",
                    input_fn=lambda p: "y", log_path=log_path,
                    collection=collection, sorted_root=sorted_root)
        assert xlsx_path.exists(), "confirmed add_invoice must create the xlsx file"
        rows = list(openpyxl.load_workbook(xlsx_path).active.values)
        assert rows[0] == ("date", "amount", "description"), "xlsx must have the header row"
        assert len(rows) == 2, f"expected 1 data row, got {len(rows) - 1}"
        assert len(Path(log_path).read_text().strip().splitlines()) == lines_before + 1, \
            "confirmed add_invoice must add exactly one undo line"

        # k. undo is scoped to the acting user (D-06): entries from two
        # different users coexist in the log; undoing as one user removes
        # only that user's entry and leaves the other user's entry in place
        write_undo_entry({
            "ts": datetime.now().isoformat(),
            "tool": "update_field",
            "codice_fiscale": cf,
            "target": "sqlite:patients.phone",
            "before": "111-1111",
            "username": "user-a",
        }, log_path)
        write_undo_entry({
            "ts": datetime.now().isoformat(),
            "tool": "update_field",
            "codice_fiscale": cf,
            "target": "sqlite:patients.phone",
            "before": "222-2222",
            "username": "user-b",
        }, log_path)

        undo_last(conn, role="dentist", username="user-b", log_path=log_path)
        assert lookup_patient(cf, conn)["phone"] == "222-2222", \
            "user-scoped undo must restore the acting user's own before-value"

        remaining_entries = [json.loads(l) for l in Path(log_path).read_text().strip().splitlines()]
        remaining_usernames = [e.get("username") for e in remaining_entries]
        assert "user-b" not in remaining_usernames, \
            "the acting user's entry must be removed after undo"
        assert "user-a" in remaining_usernames, \
            "another user's entry must remain in the log after a user-scoped undo"

    print("selftest passed")


def main():
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    positional = [a for a in sys.argv[1:] if not a.startswith("--")]

    if "--selftest" in flags:
        selftest()
        return
    if positional and positional[0] == "undo":
        session = read_session()
        if session is None:
            print("not logged in - run: python cli_session.py login")
            sys.exit(1)
        conn = init_db(DB_PATH)
        collection = get_collection("db/chroma")
        status, message = undo_last(conn, session["role"], session["username"],
                                    collection=collection, sorted_root=Path("sorted"))
        print(message)
        return
    if not positional:
        print('usage: python agent.py "<command>" [--dry-run]  |  python agent.py undo  |  python agent.py --selftest')
        sys.exit(1)

    session = read_session()
    if session is None:
        print("not logged in - run: python cli_session.py login")
        sys.exit(1)

    dry_run = "--dry-run" in flags
    conn = init_db(DB_PATH)
    collection = get_collection("db/chroma")
    try:
        run_command(positional[0], conn, dry_run, urllib.request.urlopen,
                     session["role"], session["username"],
                     collection=collection, sorted_root=Path("sorted"))
    except OllamaUnreachable as e:
        print(e)
        sys.exit(1)
    except ValueError as e:
        print("rejected:", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
